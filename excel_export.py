from openpyxl import Workbook
from db_estoque import (
    listar_produtos,
    listar_movimentacoes_produto,
    listar_produtos_baixo
)


def exportar_produtos_excel(nome_arquivo="relatorio_produtos.xlsx"):
    rows = listar_produtos()
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"

    ws.append([
        "ID", "Data", "Descricao", "Unidade", "Qtde Inicial", "Qtde Atual",
        "Fornecedor", "Estoque Minimo", "Localizacao", "Observacao"
    ])

    for r in rows:
        ws.append([
            r["id"],
            r["data_cadastro"],
            r["descricao"],
            r["unidade"],
            r["qtde_inicial"],
            r["qtde_atual"],
            r["fornecedor"],
            r["estoque_minimo"],
            r["localizacao"],
            r["observacao"]
        ])

    wb.save(nome_arquivo)


def exportar_movimentacoes_produto_excel(nome_arquivo="relatorio_movimentacoes_produto.xlsx"):
    rows = listar_movimentacoes_produto()
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimentacoes Produto"

    ws.append([
        "ID Mov", "Produto ID", "Descricao", "Unidade", "Fornecedor",
        "Tipo", "Minimo", "Saldo", "Total", "Ultima Compra",
        "Colaborador", "Usuario", "Data", "Observacao"
    ])

    for r in rows:
        qtde_atual = float(r["qtde_atual"] or 0)
        estoque_minimo = float(r["estoque_minimo"] or 0)
        saldo = qtde_atual - estoque_minimo
        total = qtde_atual
        ultima_compra = float(r["qtde_inicial"] or 0)

        ws.append([
            r["id"],
            r["produto_id"],
            r["descricao"],
            r["unidade"],
            r["fornecedor"],
            r["tipo"],
            estoque_minimo,
            saldo,
            total,
            ultima_compra,
            r["colaborador"],
            r["usuario"],
            r["data_movimentacao"],
            r["observacao"]
        ])

    wb.save(nome_arquivo)


def exportar_produtos_baixo_excel(nome_arquivo="relatorio_produtos_baixo.xlsx"):
    rows = listar_produtos_baixo()
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos Baixo"

    ws.append([
        "ID", "Descricao", "Qtde Inicial", "Qtde Atual", "Estoque Minimo", "Fornecedor"
    ])

    for r in rows:
        ws.append([
            r["id"],
            r["descricao"],
            r["qtde_inicial"],
            r["qtde_atual"],
            r["estoque_minimo"],
            r["fornecedor"]
        ])

    wb.save(nome_arquivo)